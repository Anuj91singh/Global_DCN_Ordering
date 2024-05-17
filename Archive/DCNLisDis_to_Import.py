import datetime
import glob
import numpy as np
import pandas as pd
import os
import csv
import re
import time
from itertools import islice
import urllib.request
import win32com.client as win32
from collections import Counter
from tkinter import messagebox, simpledialog,scrolledtext
import tkinter as tk
import sys
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
import openpyxl
import tkinter.font as tkFont
from openpyxl import Workbook, load_workbook
from PyPDF2 import PdfReader
from pandas.core.common import flatten
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from selenium import webdriver
from datetime import datetime, timedelta
from PyQt5.QtWidgets import QApplication, QMessageBox
import win32com.client as win32
from selenium.common.exceptions import (NoSuchElementException,
                                        NoSuchFrameException,
                                        WebDriverException)
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager

TS = datetime.now().strftime("%d, %m, %Y, %H, %M, %S")
TS = TS.split(', ')
TS_date = TS[:3]
TS_date = '-'.join(TS_date)
TS_time = TS[3:]
TS_time = '-'.join(TS_time)
date_time = '{}_{}'.format(TS_date, TS_time)

user = os.getlogin() #System login user id
download_path = (r"C:\Python_SPI\Global_DCN_test\downloads")
Query_Files_path=(r"C:\Python_SPI\Global_DCN_test\Query_Files")
files_to_delete = [f for f in os.listdir(download_path)]#deleting all existing  files in the downloads folder
for file in files_to_delete:
    os.remove(os.path.join(download_path, file))



class StatusUpdater:
    def __init__(self, root):
        self.root = root
    
        # Left side: Function names
        self.functions_var = tk.StringVar()
        bold_font = tkFont.Font(family="Times New Roman", size=16, weight="bold")
        self.functions_label = tk.Label(root, textvariable=self.functions_var, font=bold_font, bg="lightblue", padx=20, pady=20)
        self.functions_label.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Right side: Status messages
        self.status_var = tk.StringVar()
        self.status_label = tk.Label(root, textvariable=self.status_var, font=("Times New Roman", 16), bg="lightgreen", padx=20, pady=20)
        self.status_label.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # Add a scrolled text widget for multiline display
        self.text_widget = scrolledtext.ScrolledText(root, wrap=tk.WORD, font=("Times New Roman", 12), height=8)
        self.text_widget.pack(fill=tk.BOTH, expand=True)

    def update_status(self, function_name, message):
        current_content = self.text_widget.get("1.0", tk.END)
        new_content = f"{function_name}: {message}\n{current_content}"
        self.text_widget.delete("1.0", tk.END)
        self.text_widget.insert(tk.END, new_content)
        self.functions_var.set(function_name)
        self.status_var.set(message)
        self.root.update()
   

def login_EDB(user_id,password):
   
    #updater.update_status("Login","Logging in")
    #########webdriver setup################
    options = webdriver.ChromeOptions()
    prefs = {"download.default_directory" : download_path}
    options.add_experimental_option("prefs",prefs)
    #options.add_argument("--headless")
    service = Service(executable_path="C:\Python_SPI\chromedriver.exe")
    driver = webdriver.Chrome(service=service, options=options)

    #############EDB Login ###########################
    driver.get("http://edb.volvo.net/edb2/index.htm")
    driver.switch_to.frame("banner")
    driver.find_element(By.ID,"action").click()
    driver.forward()
    cred1 = driver.find_element(By.NAME, "username")
    cred1.clear()
    cred1.send_keys(user_id)
    cred2 = driver.find_element(By.NAME, "password")
    cred2.clear()
    cred2.send_keys(password)
    driver.find_element(By.CLASS_NAME, "button").click()
    driver.forward()
    driver.maximize_window()
    driver.switch_to.frame("menu")
    #updater.update_status("Login","Logged in successfully")
    return(driver)

#-----------------------------STEP 1 BEGIN------------------------------------------------------------------
def DCNDisTime(driver, current_week, updater, first_iteration=0):
    updater.update_status("DCNDisTime","Running")
    driver.find_element(By.PARTIAL_LINK_TEXT, "KOLA+").click()
    driver.find_element(By.PARTIAL_LINK_TEXT, "DCN & Object").click()
 
    WebDriverWait(driver,10).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "DCN Time from DIS"))).click()
    driver.switch_to.parent_frame()
    driver.switch_to.frame("edb_main")
    time.sleep(2)
    driver.find_element(By.ID, "time").send_keys(current_week)
    time.sleep(2)
    driver.find_element(By.XPATH, "//input[@value='Search']").click()
    #################extact the data and store in excel and convert as input ######################################
    table = driver.find_element(By.ID,'editTable') # Find the table element by its ID, XPath, or other attributes
    rows = table.find_elements(By.TAG_NAME,'tr') # Find all rows in the table body
    DCN_File = [] # Extract data from the table and store it in a list of lists
    for row in rows:
        cells = row.find_elements(By.TAG_NAME,'td')
        row_data = [cell.text.strip() for cell in cells]
        DCN_File.append(row_data)
    df = pd.DataFrame(DCN_File, columns=['', 'DCN', 'Product class', 'Factory', 'Time (YYYYWW)']) # Create a DataFrame from the extracted data

    print(df)
    DCN_File_path = os.path.join(download_path, 'DCN_File.xlsx')
    df.to_excel(DCN_File_path, index=False)# Save the DataFrame to the specific path
    wrkbk = openpyxl.load_workbook(DCN_File_path)
    sheet = wrkbk.active
    DCN_column=[sheet.cell(row=row, column=2).value for row in range(3, sheet.max_row + 1)]
    #########################################################################
    DCNfromDis_path=r'C:\Python_SPI\Global_DCN_Test\DCNfromDis.xlsx'   
    if os.path.exists(DCNfromDis_path):
            workbook = openpyxl.load_workbook(DCNfromDis_path)
            worksheet = workbook.active    
    else:
        # Step 1: Create workbook named "DCNfromDis" and name the sheet the same
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            

    df1=(df.iloc[:, 1:])
    df1 = df1.dropna(axis=0, how='all')

    existing_rows = worksheet.max_row
    for i, row in enumerate(df1.values, start=1):
            for j, value in enumerate(row, start=1):
                worksheet.cell(row=existing_rows + i, column=j, value=value)
    existing_rows = worksheet.max_row
    workbook.save(DCNfromDis_path)


    
      
    
    ###################paste the data in Kola partno info , fetch the result and downlaod csv file#############
    driver.switch_to.parent_frame()
    driver.switch_to.frame("menu")
    driver.find_element(By.PARTIAL_LINK_TEXT,'Partno Data').click()
    driver.find_element(By.PARTIAL_LINK_TEXT,'KOLA Partno Info').click()
    driver.switch_to.parent_frame()
    driver.switch_to.frame("edb_main")
    temp=Select(driver.find_element(By.NAME,'func'))
    temp.select_by_value("110")
    textarea=WebDriverWait(driver,10).until(EC.presence_of_element_located((By.XPATH,"//textarea[@name='art']")))
    textarea.click()
    for i in DCN_column:
        textarea.send_keys(str(i) + '\n')
    time.sleep(2)
    driver.find_element(By.XPATH,"//input[@name='Fetch']").click()
    driver.find_element(By.LINK_TEXT,'DOWNLOAD EXCEL FILE').click()
    time.sleep(2)
    csv_files = glob.glob(os.path.join(download_path, '*.csv'))
    if not csv_files:
        print("No CSV files found in the directory.")
    else:
        EDB123_path = (r"C:\Python_SPI\Global_DCN_Test\Query_files\EDB123.xlsx")
        if os.path.exists(EDB123_path):
            workbook = openpyxl.load_workbook(EDB123_path)
            worksheet = workbook.active
        
        else:
        # Step 1: Create workbook named "EDB123" and name the sheet the same
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            workbook.title = "EDB123"

        # Step 1.1: Add column names for "EDB123"
            column_names = ["DCN", "Type", "Heading", "Object", "Object Intro Week", "Status", "Product Class", "Factory", "DIS Time"]
            for col_index, column_name in enumerate(column_names, start=1):
                cell = worksheet.cell(row=1, column=col_index, value=column_name)
                cell.font = Font(bold=True)
            workbook.save(EDB123_path)
            
    
        latest_csv_file = max(csv_files, key=os.path.getctime)  
        with open(latest_csv_file, 'r', newline='') as file:
            reader = csv.reader(file, delimiter=';')
            # data = list(islice(reader, 1, None))
            next(reader, None)
            next(reader, None)
            first_row = next(reader,None)
            if first_row == ['DCN', 'Type', 'Heading', 'Object', 'Object Intro Week', 'Status', 'PC', 'Factory', 'DCN Intro Week']:
                data = list(reader)
            else:
                file.seek(0)
                next(reader, None) 
                next(reader, None) 
                data = list(reader)
            # data = [row for row in reader]  
         # Step 2: Ignore the first column
        m_data = [row if len(row) >= 10 else row + [''] * (10 - len(row)) for row in data]
        m_data = [row[1:] for row in m_data]
        # print("data",data)
        for row in m_data:
            if len(row)==9:
                continue
            else:
                # print("length",len(row))
                # print("row",row)             
                row[8] = row[8].replace(',', ', ')
                row[8] = row[8].replace('000000 ', '')
                row[8] = row[8].replace('000000', '')
                # print("R3",row[8])
                existing_rows = worksheet.max_row
                worksheet.append(row)
        existing_rows = worksheet.max_row
        for i, row in enumerate(m_data, start=1):
            for j, value in enumerate(row, start=1):
                worksheet.cell(row=existing_rows + i, column=j, value=value)
        
        workbook.save(EDB123_path)
    driver.quit()
    updater.update_status("DCNDisTime","completed")
    return()



def DCNListDis(updater):
        updater.update_status("DCNListDis","Running")
             
        ####now appending EDB123 data to DCNfrmKola##############
        DCNfrmkola_path=r'C:\Python_SPI\Global_DCN_Test\DCNfrmKola.xlsx'

        DCNfrmKola_df=pd.read_excel(DCNfrmkola_path)
        EDB123_path = (r"C:\Python_SPI\Global_DCN_Test\Query_files\EDB123.xlsx")
       
        EDB123_df=pd.read_excel(EDB123_path)
        EDB123_df['DIS Time'] = EDB123_df['DIS Time'].apply(lambda x: str(x))  # Convert to string
            # EDB123_df['DIS Time'] = EDB123_df['DIS Time'].apply(lambda x: ', '.join(str(int(i)) for i in x.split(', ')))
        print(EDB123_df)

        
        # Change data type of selected columns in EDB123_df to object
        selected_columns = ['Type', 'Heading', 'Object', 'Object Intro Week', 'Status', 'Product Class', 'Factory', 'DIS Time']
        for column in selected_columns:
            DCNfrmKola_df[column] = DCNfrmKola_df[column].astype(object)
        EDB123_df[selected_columns] = EDB123_df[selected_columns].apply(lambda x: x.astype(object))


         # Increment the ID column starting from the maximum value
        # Add an 'ID' column to EDB123_df
        EDB123_df.insert(0, 'ID', range(1, len(EDB123_df) + 1))
        max_id = DCNfrmKola_df['ID'].max()
        #   
        EDB123_df['ID']=range(max_id + 1, max_id + 1 + len(EDB123_df)) 
        # print(EDB123_df['ID'])
        # print(type(EDB123_df))
        # print(type(DCNfrmKola_df))
        common_columns = DCNfrmKola_df.columns.intersection(EDB123_df.columns)
        # print(common_columns)
        EDB123_df = EDB123_df[common_columns]
        # Concatenate DataFrames        
        result_df = pd.concat([DCNfrmKola_df, EDB123_df], axis=0, ignore_index=True) 
        print(result_df)
        result_df.to_excel(DCNfrmkola_path, index=False)

        ###########qrystr############
        DCNfromDis_path=r'C:\Python_SPI\Global_DCN_Test\DCNfromDis.xlsx'
        DCNOrdered_path=r'C:\Python_SPI\Global_DCN_Test\DCNOrdered.xlsx'
        DCNfromDis_df=pd.read_excel(DCNfromDis_path)
        DCNfromDis_df['DCNG'] = 'K-' + DCNfromDis_df['DCN'].str[1:6] + '-' + DCNfromDis_df['DCN'].str[6:]
        if 'DCNOrdered' not in DCNfromDis_df.columns:
            DCNfromDis_df['DCNOrdered']=None
        DCNfromDis_df.rename(columns={'Time (YYYYWW)': 'Time'}, inplace=True)
        DCNfromDis_df.to_excel(DCNfromDis_path , index=False)
        #_--------------------------------------------------------
        DCNfromDis_df=pd.read_excel(DCNfromDis_path)   
        # print(DCNfromDis_df)
        DCNOrdered_df=pd.read_excel(DCNOrdered_path, sheet_name='DCNOrdered')
        # print("DCNOrdered_df",DCNOrdered_df)
        # Merge DataFrames based on the common column DCNG and DCN Number
        merged_df = pd.merge(DCNfromDis_df, DCNOrdered_df, left_on='DCNG', right_on='DCN Number', how='left')
        print("merged_df",merged_df)
        # Update the 'DCN Ordered' column to 'Yes' where the condition is met
        merged_df['DCN Ordered'] = ''
        merged_df.loc[merged_df['DCN Number'].notnull(), 'DCN Ordered'] = 'Yes'
        result_columns =['DCN',	'Product class','Factory','Time','DCNG','DCNOrdered']
        result_df = merged_df[result_columns]
        # print("result_df",result_df)

        # current_date = pd.Timestamp.now().strftime('%Y%m%d')
        # DCNfromDis_path = f'C:\Python_SPI\Global_DCN_Test\DCNfromDis {current_date} .xlsx'
        result_df.to_excel(DCNfromDis_path , index=False)

        #############################Starting DISLISFRMDIS####################
        latest_file = max([f for f in os.listdir(r'C:\Python_SPI\Global_DCN_Test') if f.startswith('DCNfromDis')], key=os.path.getctime)
        DCNfromDis_path = os.path.join(r'C:\Python_SPI\Global_DCN_Test', latest_file)
        DCNfromDis_df = pd.read_excel(DCNfromDis_path,keep_default_na=False)

        # print("DCNfromDis_df",DCNfromDis_df)
        AppPC_path = r'C:\Python_SPI\Global_DCN_Test\Query_Files\AppPC.xlsx'
        AppPC_df = pd.read_excel(AppPC_path, sheet_name='AppPC')
        
        AppPC_df['PC'] = AppPC_df['PC'].astype(str)
        DCNfromDis_df['Product class']= DCNfromDis_df['Product class'].astype(str)
        print(DCNfromDis_df)
        print("AppPC_df",AppPC_df)
    
        

        # print("AppPC_df",AppPC_df)
        m_df = DCNfromDis_df.merge( AppPC_df , left_on='Product class', right_on='PC')
        print(m_df)

        result_df=m_df.loc[m_df['DCNOrdered'].isnull() | (m_df['DCNOrdered'] == ''), ['DCN']]
        DCNListfrmDis_path = f'C:\Python_SPI\Global_DCN_Test\downloads\DCNListfrmDis {date_time} .xlsx'
        result_df.to_excel(DCNListfrmDis_path  , index=False)

        print(f'DCNListfrmDis Completed. Results saved to: {DCNListfrmDis_path}')
        updater.update_status("DCNListDis", "Completed")
        return()
#------------------------------------STEP 1 END----------------------------------------------------------------
#--------------------------------------STEP 2 BEGIN

def Import(updater):
# def Import():
    updater.update_status("DCN is imported")  
    DCNfrmkola_path = r'C:\Python_SPI\Global_DCN_Test\DCNfrmKola.xlsx'
    Kola_df = pd.read_excel(DCNfrmkola_path)
    columns_to_null = ['Duplicate', 'DCNOrdered', 'Fact Not App', 'DCNinJobQ', 'DCNBySCP', 'PPLNotSigned', 'AMObjects']
    Kola_df.loc[:, columns_to_null] = np.NaN
    Kola_df.to_excel(DCNfrmkola_path, index=False)
    #####creating UniqueKDCN table#########
    UniqueKDCN_df=Kola_df.groupby('DCN')['ID'].max().reset_index()
    UniqueKDCN_df.rename(columns={'ID': 'MaxOfID'}, inplace=True)##need every time
    UniqueKDCN_Path=r'C:\Python_SPI\Global_DCN_Test\Query_Files\UniqueKDCN.xlsx'
    UniqueKDCN_df.to_excel(UniqueKDCN_Path, index=False)
    ########last qstr######
    UniqueKDCN_df=pd.read_excel(UniqueKDCN_Path)
    condition = (Kola_df['ID'].isin(UniqueKDCN_df['MaxOfID']))
    Kola_df.loc[condition, 'Duplicate'] = 'Unique'
    Kola_df.to_excel(DCNfrmkola_path, index=False)
    updater.update_status("DCN is imported")
    return()

#-------------------------------STEP 2 END-------------------------





