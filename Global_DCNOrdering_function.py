import glob
import numpy as np
import pandas as pd
import os
import csv
import time
from itertools import islice
from collections import Counter
from tkinter import messagebox, simpledialog,scrolledtext
import tkinter as tk
import sys
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
import openpyxl
import tkinter.font as tkFont
from openpyxl import Workbook, load_workbook
from PyPDF2 import PdfReader
from pandas.core.common import flatten
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from selenium import webdriver
from datetime import datetime
from PyQt5.QtWidgets import QApplication, QMessageBox
import win32com.client as win32
from selenium.common.exceptions import (NoSuchElementException,
                                        NoSuchFrameException,
                                        WebDriverException,TimeoutException)
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager

user = os.getlogin() #System login user id
download_path = (r"C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Temp_downloads")
Temp_join_tables_path=r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Temp_Join_Tables'
Query_Files_path=(r"C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Query_Files")
files_to_delete = [f for f in os.listdir(download_path)]#deleting all existing  files in the downloads folder
for file in files_to_delete:
    os.remove(os.path.join(download_path, file))  
files_to_delete=[f for f in os.listdir(Temp_join_tables_path)]#deleting all existing  files in the Temp_join_tables folder
for file in files_to_delete:
    os.remove(os.path.join(Temp_join_tables_path, file)) 
EDB123_path = (r"C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Query_files\EDB123.xlsx")
w=openpyxl.load_workbook(EDB123_path)
s = w.active
s.delete_rows(2, s.max_row+1)
w.save(EDB123_path)

############# locating excel ###############
Output_path=r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Output'
folder_path=r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering'
DCNfrmkola_path = r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\DCNfrmKola.xlsx'
DCNinJobQ_path = r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\DCNinJobQ.xlsx'
DCNOrdered_path=r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\DCNOrdered.xlsx'
ProjectByIH_path=r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Query_Files\ProjectByIH.xlsx'
AMObjects_path=r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Query_Files\AMObjects.xlsx'
UniqueKDCN_Path=r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Query_Files\UniqueKDCN.xlsx'
# EDB123_path = (r"C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Query_files\EDB123.xlsx")
DCNfromDis_path=r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\DCNfromDis.xlsx'
AppPC_path = r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Query_files\AppPC.xlsx'
ApplicablePSU_path= r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Query_Files\ApplicablePSU.xlsx'
ObjPrfToExcDDCN_path=r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Query_Files\ObjPrfToExcDDCN.xlsx'
  
########converting into  Dataframes########
# df_kola = pd.read_excel(DCNfrmkola_path)
# df_jobq = pd.read_excel(DCNinJobQ_path, sheet_name='DCNinJobQ')
# df_DCNOrdered=pd.read_excel(DCNOrdered_path, sheet_name='DCNOrdered')
# df_ProjectByIH=pd.read_excel(ProjectByIH_path, sheet_name='ProjectByIH')
# df_AMObjects=pd.read_excel(AMObjects_path, sheet_name='AMObjects')
# df_app_pc = pd.read_excel(AppPC_path, sheet_name='AppPC')
# df_ApplicablePSU=pd.read_excel(ApplicablePSU_path, sheet_name='ApplicablePSU')
# df_ObjPrfToExcDDCN=pd.read_excel(ObjPrfToExcDDCN_path,sheet_name='ObjPrfToExcDDCN')
today = datetime.now() 

# Set day month year format
today_str = today.strftime("%d-%m-%Y") 

class StatusUpdater:
    def __init__(self, root):
        self.root = root
    
        # Left side: Function names
        self.functions_var = tk.StringVar()
        bold_font = tkFont.Font(family="Times New Roman", size=16, weight="bold")
        self.functions_label = tk.Label(root, textvariable=self.functions_var, font=bold_font, bg="lightblue", padx=20, pady=20)
        self.functions_label.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Right side: Status messages
        self.status_var = tk.StringVar()
        self.status_label = tk.Label(root, textvariable=self.status_var, font=("Times New Roman", 16), bg="lightgreen", padx=20, pady=20)
        self.status_label.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # Add a scrolled text widget for multiline display
        self.text_widget = scrolledtext.ScrolledText(root, wrap=tk.WORD, font=("Times New Roman", 12), height=8)
        self.text_widget.pack(fill=tk.BOTH, expand=True)

    def update_status(self, function_name, message):
        current_content = self.text_widget.get("1.0", tk.END)
        new_content = f"{function_name}: {message}\n{current_content}"
        self.text_widget.delete("1.0", tk.END)
        self.text_widget.insert(tk.END, new_content)
        self.functions_var.set(function_name)
        self.status_var.set(message)
        self.root.update()
   
def login_EDB(user_id,password):
   
#    #########webdriver setup################
    options = webdriver.ChromeOptions()
    prefs = {"download.default_directory" : download_path}
    options.add_experimental_option("prefs",prefs)
    # options.add_argument("--headless")
    service = Service(executable_path="C:\Python_SPI\chromedriver.exe")
    driver = webdriver.Chrome(service=service, options=options)

    ############# EDB Login ###########################
    driver.get("http://edb.volvo.net/edb2/index.htm")
    driver.switch_to.frame("banner")
    driver.find_element(By.ID,"action").click()
    driver.forward()
    cred1 = driver.find_element(By.NAME, "username")
    cred1.clear()
    cred1.send_keys(user_id)
    cred2 = driver.find_element(By.NAME, "password")
    cred2.clear()
    cred2.send_keys(password)
    driver.find_element(By.CLASS_NAME, "button").click()
    driver.forward()
    driver.maximize_window()
    driver.switch_to.frame("menu")
    return(driver)

# # ##-----------------------------STEP 1 BEGIN--------------------------------------------
def DCNDisTime(driver, cw, updater,EDB123_path):
    updater.update_status("DCNDisTime","Running")
    driver.find_element(By.PARTIAL_LINK_TEXT, "KOLA+").click()
    driver.find_element(By.PARTIAL_LINK_TEXT, "DCN & Object").click()
 
    WebDriverWait(driver,10).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "DCN Time from DIS"))).click()
    driver.switch_to.parent_frame()
    driver.switch_to.frame("edb_main")
    time.sleep(2)
    select_element = driver.find_element(By.NAME,"op_fact")
    select = Select(select_element)
    select.select_by_value("!=")
    driver.find_element(By.XPATH,'//input[@id="fact"]').send_keys('80')
    driver.find_element(By.ID, "time").send_keys(cw)
    time.sleep(2)
    driver.find_element(By.XPATH, "//input[@value='Search']").click()
    #################extact the data and store in excel and convert as input ######################################
    table = driver.find_element(By.ID,'editTable') # Find the table element by its ID, XPath, or other attributes
    try:
        rows = table.find_elements(By.TAG_NAME,'tr') # Find all rows in the table body
    except:
        no_data=driver.find_element(By.CSS_SELECTOR,'font[color="red"]').text
        if no_data=='No data found':
            return
    DCN_File = [] # Extract data from the table and store it in a list of lists
    for row in rows:
        
        cells = row.find_elements(By.TAG_NAME,'td')
        row_data = [cell.text.strip() for cell in cells]
        if row_data is None:
            continue
        DCN_File.append(row_data)
    df = pd.DataFrame(DCN_File, columns=['', 'DCN', 'Product class', 'Factory', 'Time (YYYYWW)']) # Create a DataFrame from the extracted data

    print(df)
    DCN_File_path = os.path.join(download_path, 'DCN_File.xlsx')
    df.to_excel(DCN_File_path, index=False)# Save the DataFrame to the specific path
    wrkbk = openpyxl.load_workbook(DCN_File_path)
    sheet = wrkbk.active
    DCN_column=[sheet.cell(row=row, column=2).value for row in range(3, sheet.max_row + 1)]
    #########################################################################   
    if os.path.exists(DCNfromDis_path):
            workbook = openpyxl.load_workbook(DCNfromDis_path)
            worksheet = workbook.active    
    else:
        # Step 1: Create workbook named "DCNfromDis" and name the sheet the same
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        column_names = ["DCN", "Product Class", "Factory", "Time"]
        for col_index, column_name in enumerate(column_names, start=1):
            cell = worksheet.cell(row=1, column=col_index, value=column_name)
            cell.font = Font(bold=True)
        workbook.save(DCNfromDis_path)
            
    df1=(df.iloc[:, 1:])
    df1 = df1.dropna(axis=0, how='all')

    existing_rows = worksheet.max_row
    for i, row in enumerate(df1.values, start=1):
        for j, value in enumerate(row, start=1):
            worksheet.cell(row=existing_rows + i, column=j, value=value)
    existing_rows = worksheet.max_row
    workbook.save(DCNfromDis_path)
      
    ###################paste the data in Kola partno info , fetch the result and downlaod csv file#############
    driver.switch_to.parent_frame()
    driver.switch_to.frame("menu")
    driver.find_element(By.PARTIAL_LINK_TEXT,'Partno Data').click()
    driver.find_element(By.PARTIAL_LINK_TEXT,'KOLA Partno Info').click()
    driver.switch_to.parent_frame()
    driver.switch_to.frame("edb_main")
    temp=Select(driver.find_element(By.NAME,'func'))
    temp.select_by_value("110")
    textarea=WebDriverWait(driver,10).until(EC.presence_of_element_located((By.XPATH,"//textarea[@name='art']")))
    textarea.click()
    for i in DCN_column:
        textarea.send_keys(str(i) + '\n')
    time.sleep(2)
    driver.find_element(By.XPATH,"//input[@name='Fetch']").click()
    #dwnload=#webDriverWait(driver,1000).until(EC.presence_of_element_located((By.LINK_TEXT,'DOWNLOAD EXCEL FILE')))
    # dwnload.click()
    driver.find_element(By.LINK_TEXT,'DOWNLOAD EXCEL FILE').click()
    time.sleep(5)
    csv_files = glob.glob(os.path.join(download_path, '*.csv'))
    if not csv_files:
        print("No CSV files found in the directory.")
    else:
        
        if os.path.exists(EDB123_path):
            workbook = openpyxl.load_workbook(EDB123_path)
            worksheet = workbook.active
        
        else:
        # Step 1: Create workbook named "EDB123" and name the sheet the same
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            workbook.title = "EDB123"

        # Step 1.1: Add column names for "EDB123"
            column_names = ["DCN", "Type", "Heading", "Object", "Object Intro Week", "Status", "Product Class", "Factory", "DIS Time"]
            for col_index, column_name in enumerate(column_names, start=1):
                cell = worksheet.cell(row=1, column=col_index, value=column_name)
                cell.font = Font(bold=True)
            workbook.save(EDB123_path)
            
        latest_csv_file = max(csv_files, key=os.path.getctime)  
        with open(latest_csv_file, 'r', newline='') as file:
            reader = csv.reader(file, delimiter=';')
            # data = list(islice(reader, 1, None))
            next(reader, None)
            next(reader, None)
            first_row = next(reader,None)
            if first_row == ['DCN', 'Type', 'Heading', 'Object', 'Object Intro Week', 'Status', 'PC', 'Factory', 'DCN Intro Week']:
                data = list(reader)
            else:
                file.seek(0)
                next(reader, None) 
                next(reader, None) 
                data = list(reader)
            # data = [row for row in reader]  
        # Step 2: Ignore the first column
        m_data = [row if len(row) >= 10 else row + [''] * (10 - len(row)) for row in data]
        m_data = [row[1:] for row in m_data]
        # print("data",data)
        for row in m_data:
            if len(row)==9:
                continue
            else:
                print("length",len(row))
                print("row",row)             
                row[8] = row[8].replace(',', ', ')
                row[8] = row[8].replace('000000, ', '')
                row[8] = row[8].replace('000000', '')
                row[8] = row[8].replace(', 000000', '')
                row[8] = row[8].replace('  ', ' ')
                row[2] = row[2].replace(' ', '')
                print("R3",row[8])
                existing_rows = worksheet.max_row
                worksheet.append(row)
                print(m_data)
        existing_rows = worksheet.max_row
        for i, row in enumerate(m_data, start=1):
            for j, value in enumerate(row, start=1):
                worksheet.cell(row=existing_rows + i, column=j, value=value)
        workbook.save(EDB123_path)
        #Removing duplicates
        print("before EDB123_path",EDB123_path)
        df3 = pd.read_excel(EDB123_path)
        print("before df3 EDB123_path",df3)
        df3.drop_duplicates(inplace=True)
        print("After EDB123_path",df3)
        df3.to_excel(EDB123_path, index=False)
    driver.quit()
    updater.update_status("DCNDisTime","completed")
    return()

def DCNListDis(updater):
    updater.update_status("DCNListDis","Running")
            
    ####now appending EDB123 data to DCNfrmKola#############
    df_kola=pd.read_excel(DCNfrmkola_path)
    
    EDB123_df=pd.read_excel(EDB123_path)
    EDB123_df['DIS Time'] = EDB123_df['DIS Time'].apply(lambda x: str(x))  # Convert to string
    # EDB123_df['DIS Time'] = EDB123_df['DIS Time'].apply(lambda x: ', '.join(str(int(i)) for i in x.split(', ')))
    print(EDB123_df)

    # Change data type of selected columns in EDB123_df to object
    selected_columns = ['DCN', 'Type', 'Heading', 'Object', 'Object Intro Week', 'Status', 'Product Class', 'Factory', 'DIS Time']
    for column in selected_columns:
        df_kola[column] = df_kola[column].astype(object)
    EDB123_df[selected_columns] = EDB123_df[selected_columns].apply(lambda x: x.astype(object))
    # Increment the ID column starting from the maximum value
    # Add an 'ID' column to EDB123_df
    EDB123_df.insert(0, 'ID', range(1, len(EDB123_df) + 1))
    max_id = df_kola['ID'].max() 
    EDB123_df['ID']=range(max_id + 1, max_id + 1 + len(EDB123_df)) 
    # print(EDB123_df['ID'])
    # print(type(EDB123_df))
    # print(type(df_kola))
    common_columns = df_kola.columns.intersection(EDB123_df.columns)
    print(common_columns)
    EDB123_df = EDB123_df[common_columns]
    # Concatenate DataFrames        
    # result_df = pd.concat([df_kola, EDB123_df], axis=0, ignore_index=True) 
    result_df = pd.concat([df_kola, EDB123_df], ignore_index=True)
    print("result_df",result_df)
    result_df.to_excel(DCNfrmkola_path, index=False)
    print("DCNfrmkola_path",DCNfrmkola_path)
    print("df_kola",df_kola)
    df_kola=pd.read_excel(DCNfrmkola_path)
    print("df_kola-updated",df_kola)

    ###########qrystr############
    DCNfromDis_path=r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\DCNfromDis.xlsx'
    df_DCNfromDis=pd.read_excel(DCNfromDis_path)
    df_DCNfromDis['DCNG'] = 'K-' + df_DCNfromDis['DCN'].str[1:6] + '-' + df_DCNfromDis['DCN'].str[6:]
    if 'DCNOrdered' not in df_DCNfromDis.columns:
        df_DCNfromDis['DCNOrdered']=None
    # df_DCNfromDis.rename(columns={'Time (YYYYWW)': 'Time'}, inplace=True)
    df_DCNfromDis.to_excel(DCNfromDis_path , index=False)
    #_--------------------------------------------------------
    df_DCNfromDis=pd.read_excel(DCNfromDis_path)   
    # print(df_DCNfromDis)
    df_DCNOrdered=pd.read_excel(DCNOrdered_path, sheet_name='DCNOrdered')
    # print("df_DCNOrdered",df_DCNOrdered)
    # Merge DataFrames based on the common column DCNG and DCN Number
    merged_df = pd.merge(df_DCNfromDis, df_DCNOrdered, left_on='DCNG', right_on='DCN Number', how='left')
    print("merged_df",merged_df)
    # Update the 'DCN Ordered' column to 'Yes' where the condition is met
    merged_df['DCN Ordered'] = ''
    merged_df.loc[merged_df['DCN Number'].notnull(), 'DCN Ordered'] = 'Yes'
    result_columns =['DCN',	'Product Class','Factory','Time','DCNG','DCNOrdered']
    result_df = merged_df[result_columns]
    # print("result_df",result_df) 
    print("before DCNfromDis_path",result_df)
    result_df.drop_duplicates(inplace=True)
    print("After DCNfromDis_path",result_df)

    # current_date = pd.Timestamp.now().strftime('%Y%m%d')
    # DCNfromDis_path = f'C:\\Users\\a323151\\Desktop\\Alten_Automation\\Global_DCN_Ordering\\DCNfromDis {current_date} .xlsx'
    result_df.to_excel(DCNfromDis_path , index=False)

    #############################Starting DISLISFRMDIS####################
    # latest_file = max([f for f in os.listdir(r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering') if f.startswith('DCNfromDis')], key=os.path.getctime)
    # DCNfromDis_path = os.path.join(r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering', latest_file)
    df_DCNfromDis = pd.read_excel(DCNfromDis_path,keep_default_na=False)

    # print("df_DCNfromDis",df_DCNfromDis)
    # AppPC_path = r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Query_Files\AppPC.xlsx'
    AppPC_df = pd.read_excel(AppPC_path, sheet_name='AppPC')
    
    AppPC_df['PC'] = AppPC_df['PC'].astype(str)
    df_DCNfromDis['Product Class']= df_DCNfromDis['Product Class'].astype(str)
    print("df_DCNfromDis",df_DCNfromDis)
    print("AppPC_df",AppPC_df)

    # print("AppPC_df",AppPC_df)
    m_df = df_DCNfromDis.merge( AppPC_df , left_on='Product Class', right_on='PC')
    print("m_df",m_df)

    result_df=m_df.loc[m_df['DCNOrdered'].isnull() | (m_df['DCNOrdered'] == ''), ['DCN']]
    DCNListfrmDis_path = f'C:\\Users\\a323151\\Desktop\\Alten_Automation\\Global_DCN_Ordering\\Temp_downloads\\DCNListfrmDis {today_str} .xlsx'
    result_df.to_excel(DCNListfrmDis_path  , index=False)

    print(f'DCNListfrmDis Completed. Results saved to: {DCNListfrmDis_path}')
    updater.update_status("DCNListDis", "Completed")
    return(df_kola)
#------------------------------------STEP 1 END----------------------------------------------------------------
download_path=(r"C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\temp_downloads")
csv_files = glob.glob(os.path.join(download_path, '*.csv'))
for file in csv_files:
    os.remove(os.path.join(download_path, file)) 
# #--------------------------------------STEP 2 BEGIN--------------------------------

def Import(updater):
# def Import():
    updater.update_status("Import","Running")  
    Kola_df = pd.read_excel(DCNfrmkola_path)
    print("Import-Kola_df",Kola_df)
    # columns_to_null = ['Duplicate', 'DCNOrdered', 'Fact Not App', 'DCNinJobQ', 'DCNBySCP', 'PPLNotSigned', 'AMObjects']
    # Kola_df.loc[:, columns_to_null] = np.NaN
    # Kola_df.to_excel(DCNfrmkola_path, index=False)
    #####creating UniqueKDCN table#########
    UniqueKDCN_df=Kola_df.groupby('DCN')['ID'].max().reset_index()
    UniqueKDCN_df.rename(columns={'ID': 'MaxOfID'}, inplace=True)##need every time
    UniqueKDCN_df.to_excel(UniqueKDCN_Path, index=False)

    print("Before Unique DCN-Kola_df",Kola_df)
    print("Before Unique DCN-DCNfrmkola_path",Kola_df)
    ########last qstr###### Remove Duplicate DCN and Update Unique for column name-Duplicate
    duplicate_DCNs = Kola_df[Kola_df.duplicated(subset=['DCN'], keep='first')]
    Kola_df.loc[duplicate_DCNs.index, 'Duplicate'] = 'Duplicate'
    Kola_df.loc[~Kola_df.index.isin(duplicate_DCNs.index), 'Duplicate'] = 'Unique'
    Kola_df.drop_duplicates(subset=['DCN'], keep='first', inplace=True)
    Kola_df.to_excel(DCNfrmkola_path, index=False)
    print("After Unique DCN-Kola_df",Kola_df)
    print("After Unique DCN-DCNfrmkola_path",Kola_df)
    updater.update_status("Import","DCN is imported")
    print("import-Kola_df-DCN",Kola_df)
    print("import-Kola_df-DCN",DCNfrmkola_path)
    ####Converting DCN from"K1234567 to K-12345-67 format####
    Kola_df['DCNG'] = Kola_df['DCN'].apply(lambda value: 'K-' + value[1:6] + '-' + value[6:8])
    Kola_df.to_excel(DCNfrmkola_path, index=False)
    print("After converting DCN to K-DCN type-Kola_df-DCN",Kola_df)
    print("After converting DCN to K-DCN type",DCNfrmkola_path)

    return(Kola_df)
print("import-Kola_df-DCN",DCNfrmkola_path)
# #-------------------------------STEP 2 END-------------------------

# #---------------------STEP 3 is refreshing,STEP 4 BEGIN----------------------------

def UpdateK_DCN(updater): 
    print("UpdateK_DCN-DCNfrmkola_path",DCNfrmkola_path)
    df_kola = pd.read_excel(DCNfrmkola_path)
    print("UpdateK_DCN-df_kola",df_kola)
    df_jobq = pd.read_excel(DCNinJobQ_path, sheet_name='DCNinJobQ')
    df_DCNOrdered=pd.read_excel(DCNOrdered_path, sheet_name='DCNOrdered')
    df_ProjectByIH=pd.read_excel(ProjectByIH_path, sheet_name='ProjectByIH')
    df_AMObjects=pd.read_excel(AMObjects_path, sheet_name='AMObjects')
    updater.update_status("UpdateK_DCN" ,"Running")
    print("UpdateK_DCNdf_kola-",df_kola)

    # Update 'DCNinJobQ' column in DCNfrmKola where the condition is True
    DCNJobQUpdate_condition = df_kola['DCNG'].isin(df_jobq['DCN Design Change Notice'])
    # df_kola['DCNinJobQ'] = df_kola['DCNinJobQ'].astype(str) #alternate way
    df_kola.loc[DCNJobQUpdate_condition, 'DCNinJobQ'] = 'Yes'
    df_kola.to_excel(DCNfrmkola_path, index=False)
    print("DCNJobQUpdate_condition",df_kola)

    # Update 'DCNOrdered' column in DCNfrmKola where the condition is True
    DCNOrderUpdate_condition=df_kola['DCNG'].isin(df_DCNOrdered['DCN Number'])
    # df_kola['DCNOrdered'] = df_kola['DCNOrdered'].astype(str)   #alternate way
    df_kola.loc[DCNOrderUpdate_condition, 'DCNOrdered'] = 'Yes'
    df_kola.to_excel(DCNfrmkola_path, index=False)
    print("DCNOrderUpdate_condition",df_kola)

    # Update 'PPLNotSigned' column in DCNfrmKola where the combined condition is True
    condition1=df_kola['DCNG'].isin(df_jobq['DCN Design Change Notice'])
    condition2 = (df_jobq['DCN Archive Date'] == "1901-01-01")
    PPLNotSigned_condtion=condition1 & condition2
    df_kola.loc[PPLNotSigned_condtion, 'PPLNotSigned'] = 'Yes'
    df_kola.to_excel(DCNfrmkola_path, index=False)
    print("PPLNotSigned_condtion",df_kola)

    #Update 'DCNBySCP' column in DCNfrmKola where the condition is True
    sub_object_values = df_ProjectByIH['Sub Object'].unique()
    sub_object_list = sub_object_values.tolist()
    ProjectByIH_condition = df_kola['Object'].isin(sub_object_list)
    df_kola.loc[ProjectByIH_condition, 'DCNBySCP'] = 'Yes'
    print("ProjectByIH_condition",df_kola)
    df_kola.to_excel(DCNfrmkola_path, index=False)

    #Update 'AMObjects' column in DCNfrmKola where the condition is True
    AMObjects_values = df_AMObjects['AMObject'].unique()
    sub_object_list = AMObjects_values.tolist()
    AMObjects_condition = df_kola['Object'].isin(sub_object_list)
    df_kola.loc[AMObjects_condition, 'AMObjects'] = 'Yes'
    df_kola.to_excel(DCNfrmkola_path, index=False)
    print("AMObjects",df_kola)
    # df_kola['DCNinJobQ'].fillna('', inplace=True)
    # df_kola.to_excel(DCNfrmkola_path, index=False)
    ##Deleting null values in Duplicate column in Kola Excel
    df_kola.dropna(subset=['Duplicate'], inplace=True)
    print('Duplicate',df_kola)
    df_kola.to_excel(DCNfrmkola_path, index=False)
    ##Deleting null value in DCNInJOQ column in Kola Excel
    # df_kola.dropna(subset=['DCNinJobQ'], inplace=True)
    df_kola = df_kola.dropna(subset=['DCNinJobQ'])
    df_kola.to_excel(DCNfrmkola_path, index=False)
    print("Final",df_kola)
    print("Final-DCNfrmKola",DCNfrmkola_path)
    updater.update_status("UpdateK_DCN" ,"Completed")
    return(DCNfrmkola_path)
print("After UpdateK_DCN-Kola_df-DCN",DCNfrmkola_path)

#----------------------------STEP 4 END----------------------------

#----------------------------STEP 5 BEGIN -------------------------
def DCN_Sign_OFF_TA(updater):
    updater.update_status("DCN_Sign_OFF_TA" ,"Running")

    # Read Excel file into a DataFrame
    df_kola = pd.read_excel(DCNfrmkola_path)
    print("DCN_Sign_OFF_TA",df_kola)
    df_jobq = pd.read_excel(DCNinJobQ_path, sheet_name='DCNinJobQ')
    print("DCN_Sign_OFF_TAdf_kola-",df_kola)

    # Perform Inner Join
    merged_df = pd.merge(df_kola, df_jobq, left_on='DCNG', right_on='DCN Design Change Notice',how="inner")

    # Apply the SQL-like conditions
    query_result = merged_df[
        (merged_df['Type'] == 'Technical Authorisation') &
        (merged_df['DCNinJobQ'] == 'Yes') &
        (merged_df['PPLNotSigned'].isnull())
    ]
    # Select specific columns
    result_columns = [
        'Product Class', 'Object', 'DCNG', 'Heading', 'Type', 'DIS Time',
        'DCNinJobQ', 'PPLNotSigned', 'DCN PSU ID', 'DCN Job Role'
    ]
    final_result = query_result[result_columns]
    # current_date = datetime.now().strftime('%Y%m%d')
    DCNSignOFFTA_path = f'C:\\Users\\a323151\\Desktop\\Alten_Automation\\Global_DCN_Ordering\\Output\\DCNSignOFFTA_{today_str}.xlsx'
    final_result.to_excel(DCNSignOFFTA_path, index=False)
    print("DCNSignOFFTA Completed")

    # Remove filtered rows from original DCNfrmKola DataFrame
    df_kola = df_kola[~df_kola['DCNG'].isin(query_result['DCNG'])]
    # Save the modified DataFrame to the original Excel file
    with pd.ExcelWriter(DCNfrmkola_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_kola.to_excel(writer, sheet_name='DCNfrmKola', index=False)
    print("DCNfrmKola updated")
    updater.update_status("DCN_Sign_OFF_TA" ,"Completed")

#------------------------------STEP 5 END----------------------------

#------------------------------STEP 6 BEGIN----------------------------

def DCN_LinkID(updater):

    df_kola = pd.read_excel(DCNfrmkola_path)
    print("DCN_LinkID",df_kola)
    df_app_pc = pd.read_excel(AppPC_path, sheet_name='AppPC')
    updater.update_status("DCN_LinkID" ,"Running")
    # Perform Inner Join
    merged_df = pd.merge(df_kola, df_app_pc, left_on='Product Class', right_on='PC',how="inner")

    # Apply the SQL-like conditions
    query_result = merged_df[
        (merged_df['DIS Time'].notnull()) &
        (merged_df['Type'] == 'Product Structure') &
        (merged_df['DCNOrdered'].isnull()) &
        (merged_df['DCNinJobQ'] == 'Yes') &
        (merged_df['DCNBySCP'].isnull()) &
        (merged_df['PPLNotSigned'].isnull()) &
        (merged_df['QJ'].isin(['QJ', 'Q-']) == False) &
        (merged_df['AMObjects'].isnull())
    ]

    # Select specific columns
    result_columns = ['Product Class', 'Object', 'DCNG', 'Heading', 'DIS Time', 'Type']
    # Select specific columns and keep only distinct 'Product Class'
    final_result = query_result[result_columns].sort_values(result_columns).groupby('Product Class').apply(lambda group: group).drop_duplicates(subset=result_columns)
    print(final_result)
    # Save the result to Excel with current date
    # current_date = pd.Timestamp.now().strftime('%Y%m%d')
    DCNListLinkID_path = f'C:\\Users\\a323151\\Desktop\\Alten_Automation\\Global_DCN_Ordering\\Output\\DCNListLinkID {today_str}.xlsx'
    final_result.to_excel(DCNListLinkID_path , index=False)

    print(f'DCN_LinkID Completed. Results saved to: {DCNListLinkID_path}')
    updater.update_status("DCN_LinkID" ,"Completed")
    return()
# #  #------------------------------STEP 6 END----------------------------

# # #-------------------------------STEP 7 BEGIN--------------------------

def Generate_LinkIdList(driver,updater):
# def generate_LinkIdList(updater,driver):
    updater.update_status("Generate_LinkIdList" ,"Running")
    driver.find_element(By.PARTIAL_LINK_TEXT, "KOLA+").click()
    driver.find_element(By.PARTIAL_LINK_TEXT,'Partno Data').click()
    driver.find_element(By.PARTIAL_LINK_TEXT,'KOLA Partno Info').click()
    driver.switch_to.parent_frame()
    driver.switch_to.frame("edb_main")
    temp=Select(driver.find_element(By.NAME,'func'))
    temp.select_by_value("135")
    textarea=WebDriverWait(driver,10).until(EC.presence_of_element_located((By.XPATH,"//textarea[@name='art']")))
    textarea.click()
    folder_path = r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Output'
    latest_file = [f for f in os.listdir(Output_path) if f.startswith('DCNListLinkID ')]
    latest_file1 = max(latest_file, key=lambda f: datetime.strptime(f.split(' ')[-1].split('.')[0], '%d-%m-%Y'), default=None)
    DCNListLinkID_path = os.path.join(folder_path, latest_file1)
    df = pd.read_excel(DCNListLinkID_path)
    #Get DCNG column values
    dcng_values = df['DCNG']
    textarea.click()
    for i in dcng_values:
        textarea.send_keys(str(i) + '\n')
    WebDriverWait(driver,4000).until(EC.presence_of_element_located((By.XPATH,"//input[@name='Fetch']"))).click()    
    # driver.find_element(By.XPATH,"//input[@name='Fetch']").click()
    try:
        WebDriverWait(driver,6000).until(EC.presence_of_element_located((By.LINK_TEXT,"DOWNLOAD EXCEL FILE"))).click()
    except TimeoutException:
        print("Timeout: The 'DOWNLOAD EXCEL FILE' link did not appear within the specified time.")
        # driver.find_element(By.LINK_TEXT,'DOWNLOAD EXCEL FILE').click()
        time.sleep(30)

    max_wait_time = 300
    start_time = time.time()
    while True:
        csv_files = glob.glob(os.path.join(download_path, '*.csv'))
        if csv_files:
            latest_csv_file = max(csv_files, key=os.path.getctime)
            if os.path.getsize(latest_csv_file) > 0:
                break
        if time.time() - start_time >= max_wait_time:
            print("Timeout: CSV file download took too long.")
            break
        time.sleep(1)
    print("latest_csv_file",latest_csv_file)
    if csv_files:
        with open(latest_csv_file, 'r', newline='') as infile:
            reader = csv.reader(infile,delimiter=';')
            next(infile)
            workbook = Workbook()
            sheet = workbook.active
            header = next(reader)
            sheet.append(header[:24])
            for row in reader:
                sheet.append(row[:24])

    workbook.save(r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Temp_downloads\Edbdata.xlsx')
    df=pd.read_excel(r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Temp_downloads\Edbdata.xlsx')
    df1=pd.read_excel(r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\DCNLinkID.xlsx')
    common_columns = df1.columns.intersection(df.columns)
    df = df[common_columns]
    # Concatenate DataFrames        
    df2 = pd.concat([df1, df], axis=0, ignore_index=True) 
    # df2=pd.concat([df1,df], axis=1) 
    print(df)
    df2.to_excel((r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\DCNLinkID.xlsx'),index=False)
    # with pd.ExcelWriter(r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\DCNLinkID.xlsx', engine='openpyxl', mode='a') as writer:
    #     df2.to_excel(writer, index=False, header=False, if_sheet_exists=replace ,startcol=df1.shape[1])
    updater.update_status("Generate_LinkIdList" ,"Completed")
    return()
##-------------------------------Step 7 END----------------------------------

##-------------------------------Step 8 BEGIN--------------------------------

def DCN_SignOffLinkID(updater):
    df_kola = pd.read_excel(DCNfrmkola_path)
    df_jobq = pd.read_excel(DCNinJobQ_path, sheet_name='DCNinJobQ')
    updater.update_status("DCN_SignOffLinkID" ,"Running")
    #First Create DCNtoOrder excel
    DCNLinkID_path=r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\DCNLinkID.xlsx'
    df_DCNLinkID=pd.read_excel(DCNLinkID_path)
    # Merge DataFrames based on the DCN column
    merged_df = pd.merge(df_DCNLinkID, df_kola, left_on='DCN', right_on='DCN')
    # Filter rows based on the Parttype column
    merged_df.columns = merged_df.columns.map(lambda x: x.rstrip('_x'))
    filtered_df = merged_df[merged_df['Parttype'].isin(['K', 'D', 'S', 'P'])]
    # Select specific columns
    result_columns = ['Product Class', 'DCNG', 'Object', 'Heading', 'KOLA Time', 'DIS Time']
    # final_result = filtered_df[result_columns].sort_values(by=result_columns).drop_duplicates()
    final_result = filtered_df[result_columns].sort_values(result_columns).groupby('Product Class').apply(lambda group: group).drop_duplicates(subset=result_columns)
    DCNtoOrder_path= f'C:\\Users\\a323151\\Desktop\\Alten_Automation\\Global_DCN_Ordering\\DCNtoOrder {today_str}.xlsx'
    final_result.to_excel(DCNtoOrder_path, index=False)
    print("DCNtoOrder created")
    #######STARTING SIGNOFFLINKID####
    folder_path = r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering'
    latest_file = max([f for f in os.listdir(folder_path) if f.startswith('DCNtoOrder')], key=os.path.getctime)
    DCNtoOrder_path = os.path.join(folder_path, latest_file)
    DCNtoOrder_df = pd.read_excel(DCNtoOrder_path)
    DCNtoOrder_df = DCNtoOrder_df.rename(columns={'DCNG': 'DCNG_y'})
    DCNtoOrder_df = DCNtoOrder_df.rename(columns={'Heading': 'Heading_y'})
    latest_file = max([f for f in os.listdir(Output_path) if f.startswith('DCNListLinkID ')], default=None)
    DCNListLinkID_path = os.path.join(Output_path, latest_file) if latest_file else None
    df_DCNListLinkID = pd.read_excel(DCNListLinkID_path)

    # Perform the SQL-like query
    temp_leftjoin_df = df_DCNListLinkID.merge(DCNtoOrder_df, left_on='DCNG',right_on='DCNG_y',how='left')
    # print("temp_leftjoin_df:", temp_leftjoin_df.columns)
    # print(temp_leftjoin_df)
    merged_df = temp_leftjoin_df.merge(df_jobq, left_on='DCNG', right_on='DCN Design Change Notice', how='inner') 
    # print("merged_df:", merged_df.columns)
    # temp_leftjoin_df.to_excel((f'C:\\Users\\a323151\\Desktop\\Alten_Automation\\Global_DCN_Ordering\\emp_leftjoin_df{date_time}.xlsx'),index=False)
    # print(merged_df)
    filtered_df = merged_df[merged_df['DCNG_y'].isna()]
    # merged_df.columns = merged_df.columns.map(lambda x: x.rstrip('_x'))
    # merged_df.columns = merged_df.columns.map(lambda x: x.rstrip('_y'))
    # print(merged_df)
    # print("filtered_df:", filtered_df.columns)
    # filtered_df = filtered_df.rename(columns={'DCNG_y': 'DCNG_y'})
    # print(filtered_df)
    # print("print(filtered_df['DCNG_y']",filtered_df['DCNG_y'])
    # Select specific columns
    result_columns = ['DCNG', 'Heading', 'DCN PSU ID', 'DCN Job Role', 'Product Class_x', 'Object_x', 'DIS Time_x']
    # Get unique values based on the specified columns
    unique_result_df = filtered_df[result_columns].sort_values(result_columns).groupby('DCNG').apply(lambda group: group).drop_duplicates(subset=result_columns)
    print("unique_result_df",unique_result_df)
    # Save the result to Excel with current date
    # current_date = pd.Timestamp.now().strftime('%Y%m%d')
    DCN_SignOffLinkID_path = f'C:\\Users\\a323151\\Desktop\\Alten_Automation\\Global_DCN_Ordering\\Output\\DCN_SignOffLinkID {today_str} .xlsx'
    unique_result_df.to_excel(DCN_SignOffLinkID_path , index=False)

    print(f'DCN_SignOffLinkID Completed. Results saved to: {DCN_SignOffLinkID_path }')
    updater.update_status("DCN_SignOffLinkID" ,"Completed")

##-----------------------------STEP 8 END---------------------------------------------------------

##-----------------------------STEP 9 BEGIN-----------------------------------------------------
    ###twp files generated here DCNtoOrderPSU and DCNLinkIDtoOrder
# def DCNtoOrderPSU(updater):
def KDCN_toOrder(updater):

    df_jobq = pd.read_excel(DCNinJobQ_path, sheet_name='DCNinJobQ')

    # updater.update_status("KDCN_toOrder" ,"Running")
    # updater.update_status("DCNtoOrderPSU" ,"Running")
    folder_path = r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering'
    latest_file = max([f for f in os.listdir(folder_path) if f.startswith('DCNtoOrder ')])
    DCNtoOrder_path = os.path.join(folder_path, latest_file)
    df_DCNtoOrder = pd.read_excel(DCNtoOrder_path) #DCNtoOrder df
    result_df = pd.merge(df_DCNtoOrder,df_jobq,how='inner',left_on='DCNG',right_on='DCN Design Change Notice'
    )[['Product Class','DCNG', 'Object', 'Heading', 'KOLA Time', 'DIS Time', 'DCN PSU ID', 'DCN PSD Site']].drop_duplicates()
    DCNtoOrderPSU_path = f'C:\\Users\\a323151\\Desktop\\Alten_Automation\\Global_DCN_Ordering\\Output\\DCNtoOrderPSU {today_str} .xlsx'
    result_df.to_excel(DCNtoOrderPSU_path,index=False)
    # updater.update_status("DCNtoOrderPSU" ,"Completed")
    print(f'DCNtoOrderPSU Completed. Results saved to: {DCNtoOrderPSU_path }')

# def DCNLinkIDtoOrder(updater):
    updater.update_status("DCNLinkIDtoOrder" ,"Running")
    df_kola=pd.read_excel(DCNfrmkola_path)
    print("KDCN_toOrder_df_kola-",df_kola)
    df_kola=df_kola.astype("object")
    DCN_LinkID_Path= r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\DCNLinkID.xlsx'
    df_DCNLinkID = pd.read_excel(DCN_LinkID_Path)
    df_DCNLinkID=df_DCNLinkID.astype("object")
    # DCNLinkIDs['Function Group'] = DCNLinkIDs['Function Group'].apply(lambda x: str(x))
    result_df = pd.merge(df_DCNLinkID[df_DCNLinkID['Parttype'].isin(['K', 'D', 'S'])],  # Filter by Parttype
                         df_kola,how='inner',on='DCN'
                         )[['Product Class_x','Function Group','Suffix','Context','Usage','DCNG','Description','Object_x','Status_x','LinkID',
                            'Partno','Name','Qty','Variants','Added','Deleted','Module','Parttype','KOLA Time_x','DIS Time']].drop_duplicates()
    DCNLinkIDtoOrder_df = result_df.rename(columns={'Product Class_x': 'Product Class','Object_x':'Object','KOLA Time_x':'KOLA Time','Status_x':'Status'})
    DCNLinkIDtoOrder_df['Function Group'] = DCNLinkIDtoOrder_df['Function Group'].astype(str)
    DCNLinkIDtoOrder_df['FG1'] = DCNLinkIDtoOrder_df['Function Group'].apply(lambda x: x[0])
    DCNLinkIDtoOrder_df['FG2'] = DCNLinkIDtoOrder_df['Function Group'].apply(lambda x: x[:2])
    DCNLinkIDtoOrder_path=f'C:\\Users\\a323151\\Desktop\\Alten_Automation\\Global_DCN_Ordering\\Output\\DCNLinkIDtoOrder {today_str}.xlsx'
    DCNLinkIDtoOrder_df.to_excel(DCNLinkIDtoOrder_path,index=False)
    updater.update_status("DCNLinkIDtoOrder" ,"Completed")
    # updater.update_status("KDCN_toOrder" ,"Completed")
    print(f'DCNLinkIDtoOrder. Results saved to: {DCNLinkIDtoOrder_path }')

##----------------------------STEP 9 END-------------------------------------
    
##----------------------------STEP 10 BEGIN----------------------------------
    ### in step 10 we need to create 2 output sheet#KDCNListQJAM and DDCNLISTQJM
def KDCNListQJAM_DDCNListQJAM(updater):
    df_kola = pd.read_excel(DCNfrmkola_path)

#creating DCNtoOrderAM
    updater.update_status("DCNListQJAM&DDCNListQJAM" ,"Running")
    df_Kola = pd.read_excel(DCNfrmkola_path)
    print("KDCNListQJAM_DDCNListQJAM_df_kola-",df_kola)
    DCNtoOrderAM_df = df_Kola.loc[
        (df_Kola['AMObjects'] == 'Yes') &
        (df_Kola['DCNOrdered'].isnull()) &
        (df_Kola['DCNinJobQ'] == 'Yes'),
        ['DCNG', 'Object', 'Heading', 'DIS Time']
    ].drop_duplicates()
    DCNtoOrderAM_path=r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Temp_Join_Tables\DCNtoOrderAM.xlsx'
    DCNtoOrderAM_df.to_excel(DCNtoOrderAM_path,index=False)
    print("DCNtoOrderAM is created")
##creating DCNtoOrderQJ
    df_Kola = pd.read_excel(DCNfrmkola_path)
    DCNtoOrderQJ = df_Kola.loc[
        ((df_Kola['QJ'] == 'QJ') | (df_Kola['QJ'] == 'Q-')) &
        (df_Kola['DCNOrdered'].isnull()) &
        (df_Kola['DCNinJobQ'] == 'Yes'),
        ['DCNG', 'Object', 'Heading', 'DIS Time', 'QJ']
    ].drop_duplicates()
    DCNtoOrderQJ_path=r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Temp_Join_Tables\DCNtoOrderQJ.xlsx'
    DCNtoOrderQJ.to_excel(DCNtoOrderQJ_path,index=False)
    print("DCNtoOrderQJ is created")
#creating KDCNListQJAM
    DCNtoOrderAM_path=r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Temp_Join_Tables\DCNtoOrderAM.xlsx'
    DCNtoOrderQJ_path=r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Temp_Join_Tables\DCNtoOrderQJ.xlsx'
    df1 = pd.read_excel(DCNtoOrderAM_path) #DCNtoOrderAM df  
    df2 = pd.read_excel(DCNtoOrderQJ_path) #DCNtoOrderQJ df
    df_KDCNListQJAM= pd.concat([df1[['DCNG', 'Object', 'Heading', 'DIS Time']],
                           df2[['DCNG', 'Object', 'Heading', 'DIS Time']]]).drop_duplicates()
    KDCNListQJAM_path=f'C:\\Users\\a323151\\Desktop\\Alten_Automation\\Global_DCN_Ordering\\Output\\KDCNListQJAM {today_str}.xlsx'
    df_KDCNListQJAM.to_excel(KDCNListQJAM_path,index=False)
    # updater.update_status("KDCNListQJAM" ,"Completed")
    print(f'KDCNListQJAM. Results saved to: {KDCNListQJAM_path}')

##creating DDCNAM
    df_AMObjects=pd.read_excel(AMObjects_path, sheet_name='AMObjects')
    df_jobq = pd.read_excel(DCNinJobQ_path, sheet_name='DCNinJobQ')
    df_DDCNAM = pd.merge(df_AMObjects,df_jobq,how='inner',left_on='AMObject',right_on='DCN Object Number'
    ).loc[df_jobq['DCN Archive Date'] != '1901-01-01',
        ['DCN Object Number', 'DCN Design Change Notice', 'DCN PSU ID']
    ].drop_duplicates()
    DDCNAM_path=r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Temp_Join_Tables\DDCNAM.xlsx'
    df_DDCNAM.to_excel(DDCNAM_path,index=False)
    print("DDCNAM is created")
##creating DDCNQJ
    df_DDCNQJ= df_jobq.loc[
        (df_jobq['DCN Archive Date'] != '1901-01-01') &
        (df_jobq['DCN Object Number'].str[:2].isin(['QJ', 'Q-'])) &
        (df_jobq['DCN Design Change Notice 1-1'] == 'D'),
        ['DCN Object Number', 'DCN Design Change Notice', 'DCN PSU ID']
    ].drop_duplicates()
    DDCNQJ_path=r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Temp_Join_Tables\DDCNQJ.xlsx'
    df_DDCNQJ.to_excel(DDCNQJ_path,index=False)
    print("DDCNQJ is created")
###DDCNListQJAM
    # updater.update_status("DDCNListQJAM" ,"Running")
    df_DDCN_AM = pd.read_excel(r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Temp_Join_Tables\DDCNAM.xlsx')
    df_DDCN_QJ = pd.read_excel(r'C:\Users\a323151\Desktop\Alten_Automation\Global_DCN_Ordering\Temp_Join_Tables\DDCNQJ.xlsx')
    df_DDCNListQJAM = pd.concat([
       df_DDCN_AM[['DCN Object Number', 'DCN Design Change Notice', 'DCN PSU ID']],
        df_DDCN_QJ[['DCN Object Number', 'DCN Design Change Notice', 'DCN PSU ID']]
    ]).drop_duplicates()
    DDCNListQJAM_path=f'C:\\Users\\a323151\\Desktop\\Alten_Automation\\Global_DCN_Ordering\\Output\\DDCNListQJAM {today_str}.xlsx'
    df_DDCNListQJAM.to_excel(DDCNListQJAM_path,index=False)
    updater.update_status("DCNListQJAM&DDCNListQJAM" ,"Completed")
    print(f'DDCNListQJAM. Results saved to: {DDCNListQJAM_path}')

####----------------------------STEP 10 END-------------------------------

###-----------------------------STEP 11 BEGIN-----------------------------
def DDCNList(updater):
    df_ApplicablePSU=pd.read_excel(ApplicablePSU_path, sheet_name='ApplicablePSU')

    ##first create DCNListingS1 using DCNinJobQ and Applicable PSU table
    updater.update_status("DDCNList" ,"Running")
    df_jobq=pd.read_excel(DCNinJobQ_path,sheet_name='DCNinJobQ')
    df_jobq['DCN PSU ID'] = pd.to_numeric(df_jobq['DCN PSU ID'], errors='coerce')
    df_ApplicablePSU['PSU'] = pd.to_numeric(df_ApplicablePSU['PSU'], errors='coerce')
    merged_df = df_jobq.merge(df_ApplicablePSU, left_on='DCN PSU ID', right_on='PSU', how='inner')
    # print(merged_df)
    ##Apply WHERE conditions
    filtered_df = merged_df[
        (merged_df['DCN Archive Date'] != "1901-01-01") &
        (merged_df['DCN Design Change Notice 1-1'] == "D") &
        ((merged_df['DCN Object Number'].str[:2] != "QJ") & (merged_df['DCN Object Number'].str[:2] != "Q-"))
    ]
    # print(filtered_df)
    result_df = pd.DataFrame({
    'DCN Design Change Notice': filtered_df['DCN Design Change Notice'],
    'DCN Object Number': filtered_df['DCN Object Number'],
    'DCN PSU ID': filtered_df['DCN PSU ID'],
    'DCN Archive Date': filtered_df['DCN Archive Date'],
    'Site': filtered_df['Site'],
    'Prefix': [x[:3] for x in filtered_df['DCN Object Number']],  # Prefix
    'Prefix1': [x[:2] for x in filtered_df['DCN Object Number']],  # Prefix1
    })
    result_columns =['DCN Design Change Notice','DCN Object Number', 'DCN PSU ID', 'DCN Archive Date', 'Site', 'Prefix', 'Prefix1']
    final_result = result_df[result_columns].sort_values(result_columns).groupby('DCN Design Change Notice').apply(lambda group: group).drop_duplicates(subset=result_columns)
    final_result['DCN PSU ID'] = final_result['DCN PSU ID'].astype('Int64')  # Convert to nullable integer
    final_result['DCN PSU ID'] = final_result['DCN PSU ID'].astype(str)  # Convert to string
    # print(final_result)
    DCNListingS1_path = f'C:\\Users\\a323151\\Desktop\\Alten_Automation\\Global_DCN_Ordering\\Temp_Join_Tables\\DCNListingS1.xlsx'
    final_result.to_excel(DCNListingS1_path , index=False)
    print("DCNListingS1 Completed.")
    #####DDCNlistings2
    df_DCNListingS1 = pd.read_excel(DCNListingS1_path)
    df_AMObjects=pd.read_excel(AMObjects_path, sheet_name='AMObjects')
    df_ObjPrfToExcDDCN=pd.read_excel(ObjPrfToExcDDCN_path,sheet_name='ObjPrfToExcDDCN')
    df_ObjPrfToExcDDCN = df_ObjPrfToExcDDCN.rename(columns={'Prefix': 'Prefix_y'})
    merged_df = df_DCNListingS1.merge(df_ObjPrfToExcDDCN, left_on='Prefix', right_on='Prefix_y', how='left',suffixes=('', '_right'))
    print(merged_df.columns)
    merged_df = pd.merge(merged_df, df_AMObjects, left_on='DCN Object Number', right_on='AMObject', how='left')
    filtered_df = merged_df[(merged_df['Prefix_y'].isna()) & (merged_df['AMObject'].isna())]
    filtered_df['ArchYear']=filtered_df['DCN Archive Date'].astype(str).str[:4]
    result_columns = ['DCN Design Change Notice', 'DCN Object Number', 'DCN PSU ID', 'Site', 'ArchYear']
    df_DDCNList = filtered_df[result_columns]
    DDCNList_Path  = f'C:\\Users\\a323151\\Desktop\\Alten_Automation\\Global_DCN_Ordering\\Output\\DDCNList {today_str}.xlsx'
    df_DDCNList.to_excel(DDCNList_Path, index=False)
    updater.update_status("DDCNList" ,"Completed")
    print(f'DDCNList created. Results saved to: {DDCNList_Path}')

##--------------------------------STEP 11 END--------------------------------

##--------------------------------STEP 13 BEGIN------------------------------
def DCNNotSignedByPPL(updater):
    updater.update_status("DCNNotSignedByPPL","Running")
    df_kola = pd.read_excel(DCNfrmkola_path)
    print("DCNNotSignedByPPL_df_kola-",df_kola)
    df_jobq= pd.read_excel(DCNinJobQ_path,sheet_name='DCNinJobQ')
    df = pd.merge(df_kola, df_jobq, how='inner', left_on='DCNG', right_on='DCN Design Change Notice')
    df = df[(df['DIS Time'].notnull()) & 
            (df['PPLNotSigned']=='Yes') &
            (df['DCN Job Role']=='PPL')]  
    df = df[['DCNG', 'Object', 'DIS Time', 'DCN PSU ID', 'DCNBySCP']].drop_duplicates()
    DCNNotSignedByPPL_path=f'C:\\Users\\a323151\\Desktop\\Alten_Automation\\Global_DCN_Ordering\\Output\\DCNNotSignedByPPL {today_str}.xlsx'
    df.to_excel(DCNNotSignedByPPL_path,index=False)
    updater.update_status("DCNNotSignedByPPL" ,"Completed")
    print(f'DCNNotSignedByPPL is created. Results saved to: {DCNNotSignedByPPL_path}')
    updater.update_status("DCN_Ordering Script" ,"successfully completed")

##-------------------------------STEP 13 END-------------------------------------

